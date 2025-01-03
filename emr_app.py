import sys
import subprocess
import requests
import zipfile
import json
import uuid
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget,
    QTableWidget, QTableWidgetItem, QMessageBox, QHBoxLayout, QCheckBox, QAction
)
from PyQt5.QtChart import QChart, QChartView, QLineSeries
from PyQt5.QtCore import Qt, QDate
import os
from PyQt5.QtWidgets import QFileDialog, QLabel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import tempfile
import logging
import os
from fpdf import FPDF  # Import the FPDF library for PDF creation
import tarfile
import stat

DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
VERSION = "v1.0.25"


def get_user_data_path(filename):
    """Get the absolute path to store user-specific resource files."""
    # Determine the base path depending on the operating system
    if os.name == "nt":  # Windows
        base_path = os.getenv("LOCALAPPDATA", os.path.expanduser("~\\AppData\\Local"))
    elif sys.platform == "darwin":  # macOS
        base_path = os.path.expanduser("~/Library/Application Support")
    else:  # Assume Linux or other Unix-based system
        base_path = os.path.expanduser("~/.config")

    # Create the application-specific directory if it doesn't exist
    app_folder = os.path.join(base_path, "CaseManager")
    os.makedirs(app_folder, exist_ok=True)

    # Return the full path for the file
    return os.path.join(app_folder, filename)


def is_new_version_on_platform2(name):
    platform_name = platform.system()
    is_version_to_update = False

    if platform_name == "Windows":
        is_version_to_update = name.split('.zip')[0].endswith("_windows")
    elif platform_name == "Darwin":  # macOS
        is_version_to_update = name.split('.tar.gz')[0].endswith("_macos")

    return is_version_to_update


def setup_logging():
    log_dir = get_user_data_path("")  # Get the CaseManager directory
    log_file = os.path.join(log_dir, "app.log")

    # Configure logging
    logging.basicConfig(
        filename=log_file,
        filemode="a",  # Append to the log file
        format="%(asctime)s - %(levelname)s - %(message)s",
        level=logging.DEBUG,
    )
    logging.debug("Logging setup complete.")


setup_logging()


class EMRManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"Case Manager {VERSION}")
        self.setGeometry(100, 100, 900, 600)

        # Main layout
        self.main_widget = QWidget()
        self.main_layout = QVBoxLayout()

        # Table to display patients
        self.patient_table = QTableWidget()
        self.patient_table.setColumnCount(3)
        self.patient_table.setHorizontalHeaderLabels(["Name", "Age", "Data"])
        self.main_layout.addWidget(self.patient_table)

        # Connect cellChanged signal to handle table edits
        self.patient_table.cellChanged.connect(self.update_patient_data)

        # Load patient data
        self.patients = self.load_patients()
        self.populate_table()

        # Buttons
        self.update_button = QPushButton("Check for updates")
        self.update_button.clicked.connect(self.check_for_updates)
        self.main_layout.addWidget(self.update_button)

        add_button = QPushButton("Add Patient")
        add_button.clicked.connect(self.add_patient)
        self.main_layout.addWidget(add_button)

        delete_button = QPushButton("Delete Patient")
        delete_button.clicked.connect(self.delete_patient)
        self.main_layout.addWidget(delete_button)

        # Set up the main widget
        self.main_widget.setLayout(self.main_layout)
        self.setCentralWidget(self.main_widget)

        # Menu for settings
        self.menu_bar = self.menuBar()
        settings_menu = self.menu_bar.addMenu("Settings")

        # Add Edit Data button to settings
        edit_data_action = settings_menu.addAction("Edit Data")
        edit_data_action.triggered.connect(self.open_edit_data_screen)

        # Load questions
        self.questions = self.load_questions()

    def check_for_updates(self):
        """Check for updates using GitHub API."""
        repo_owner = "marcinknara"
        repo_name = "minimal-emr"
        current_version = self.get_current_version()

        try:
            # Fetch the latest release data from GitHub API
            url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/releases/latest"
            response = requests.get(url)
            response.raise_for_status()
            release_data = response.json()

            # download_url = release_data['zipball_url']
            download_url = ''
            latest_version = release_data["tag_name"]
            for release in release_data["assets"]:
                if is_new_version_on_platform2(release['name']):
                    download_url = release["browser_download_url"] if release["browser_download_url"] else None

            if latest_version > current_version and download_url:
                logging.info(f"Download URL: {download_url}")

                choice = QMessageBox.question(
                    self,
                    "Update Available",
                    f"A new version ({latest_version}) is available. Current version({current_version}). Do you want to update?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if choice == QMessageBox.Yes and download_url:
                    self.download_and_apply_update(download_url, latest_version)
                else:
                    QMessageBox.information(self, "Update", "Update canceled.")
            else:
                QMessageBox.information(self, "No Updates", "You are using the latest version.")

        except requests.RequestException as e:
            QMessageBox.critical(self, "Error", f"Failed to check for updates: {e}")

    def download_and_apply_update(self, download_url, latest_version):
        """Download and apply the update."""
        try:
            temp_dir = tempfile.mkdtemp()
            if platform.system() == 'Windows':
                archive_path = os.path.join(temp_dir, "update.zip")
            else:  # macOS
                archive_path = os.path.join(temp_dir, "update.tar.gz")

            # Download the update
            response = requests.get(download_url, stream=True)
            with open(archive_path, "wb") as update_file:
                for chunk in response.iter_content(chunk_size=1024):
                    if chunk:
                        update_file.write(chunk)

            # Extract the update
            if platform.system() == 'Windows':
                with zipfile.ZipFile(archive_path, "r") as zip_ref:
                    zip_ref.extractall(temp_dir)
            else:  # macOS
                with tarfile.open(archive_path, "r:gz") as tar:
                    tar.extractall(temp_dir)

            # Apply the update
            self.apply_update(temp_dir, latest_version)

        except requests.RequestException as e:
            QMessageBox.critical(self, "Error", f"Failed to download update: {e}")
        except zipfile.BadZipFile:
            QMessageBox.critical(self, "Error", "Downloaded update file is corrupted.")

    def apply_update(self, update_dir, latest_version):
        """Replace the current executable with the new one and restart."""
        import shutil
        try:
            # Path to the current running executable
            current_exe = sys.executable
            logging.info(f"Current executable path: {current_exe}")

            # Extract update.zip
            for filename in os.listdir(update_dir):
                file_path = os.path.join(update_dir, filename)
                if platform.system() == 'Windows' and filename == 'update.zip':
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        zip_ref.extractall(update_dir)
                elif platform.system() == 'Darwin' and filename == 'update.tar.gz':
                    with tarfile.open(file_path, 'r:gz') as tar:
                        tar.extractall(update_dir)

            # Find the new executable (assume it's inside the extracted folder)
            subdirs = [d for d in os.listdir(update_dir) if os.path.isdir(os.path.join(update_dir, d))]
            random_dir = os.path.join(update_dir, subdirs[0])

            new_exe_path = None
            for root, _, files in os.walk(random_dir):
                for file in files:
                    if platform.system() == 'Windows':
                        if file.endswith('.exe'):  # Assume the new exe has .exe extension
                            new_exe_path = os.path.join(root, file)
                            break
                    elif platform.system() == 'Darwin':
                        if file == 'CaseManager':
                            new_exe_path = os.path.join(root, file)

                            st = os.stat(new_exe_path)
                            os.chmod(new_exe_path, st.st_mode | stat.S_IEXEC)
                            break

            if not new_exe_path:
                raise FileNotFoundError("New executable not found in the update package.")

            logging.info(f"New executable path: {new_exe_path}")
            if platform.system() == 'Windows':
                # Replace current executable with the new one
                backup_exe = current_exe + ".bak"
                if os.path.exists(backup_exe):
                    os.remove(backup_exe)

                os.rename(current_exe, backup_exe)  # Backup current exe
                shutil.move(new_exe_path, current_exe)

                # Update version.json
                existing_dir = os.getcwd()
                app_dir = os.path.abspath(os.path.join(existing_dir, ".."))
                with open(os.path.join(app_dir, "version.json"), "w") as version_file:
                    json.dump({"version": latest_version}, version_file, indent=4)

                # Restart the application
                QMessageBox.information(self, "Update", "Update applied successfully. Restarting...")
                subprocess.Popen([current_exe])
                sys.exit(0)
            else:
                # Backup the current application
                backup_app_path = current_exe + "_backup"
                if os.path.exists(backup_app_path):
                    if os.path.isfile(backup_app_path):
                        os.remove(backup_app_path)
                    elif os.path.isdir(backup_app_path):
                        shutil.rmtree(backup_app_path)

                shutil.move(current_exe, backup_app_path)

                # Replace the current application with the updated version
                shutil.move(new_exe_path, current_exe)

                # Update version.json (optional)
                app_dir = os.path.dirname(current_exe)
                with open(os.path.join(app_dir, "version.json"), "w") as version_file:
                    json.dump({"version": latest_version}, version_file, indent=4)

                # Restart the updated application
                QMessageBox.information(self, "Update", "Update applied successfully. Restarting...")
                subprocess.Popen(["open", current_exe])  # macOS command to open the app
                sys.exit(0)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to apply update: {e}")
            logging.error(f"Update failed: {e}")

    def get_current_version(self):
        """Read the current version from version.json."""
        try:
            current_dir = os.getcwd()
            parent_directory = os.path.abspath(os.path.join(current_dir, ".."))
            version_file_dir = os.path.join(parent_directory, 'version.json')

            if platform.system() == 'Darwin':
                current_exe = sys.executable
                app_dir = os.path.dirname(current_exe)
                version_file_dir = os.path.join(app_dir, 'version.json')

            with open(version_file_dir, "r") as version_file:
                version_data = json.load(version_file)
                return version_data.get("version", "0.0.0")
        except (FileNotFoundError, json.JSONDecodeError):
            # QMessageBox.warning(self, "Error", "Version file not found or corrupted. Assuming version 0.0.0.")
            return "0.0.0"

    def load_patients(self):
        try:
            path = get_user_data_path("patients.json")
            logging.info(f"Loading patients from: {path}")  # Debugging information
            if not os.path.exists(path):
                with open(path, "w") as file:
                    json.dump({}, file, indent=4)
                    logging.info("Created new patients.json file.")  # Debugging information
            with open(path, "r") as file:
                return json.load(file)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            QMessageBox.warning(self, "Error", f"Failed to load patients: {str(e)}")
            return {}

    def save_patients(self):
        try:
            with open(get_user_data_path("patients.json"), "w") as file:
                json.dump(self.patients, file, indent=4)
        except IOError as e:
            QMessageBox.critical(self, "Error", f"Failed to save patients: {str(e)}")

    def populate_table(self):
        """Populate the patient table."""
        self.patient_table.blockSignals(True)
        self.patient_table.setRowCount(0)
        for uuid_key, patient in self.patients.items():
            try:
                row = self.patient_table.rowCount()
                self.patient_table.insertRow(row)

                name_item = QTableWidgetItem(patient.get("name", "Unnamed Patient"))
                name_item.setFlags(name_item.flags() | Qt.ItemIsEditable)
                self.patient_table.setItem(row, 0, name_item)

                age_item = QTableWidgetItem(str(patient.get("age", 30)))
                age_item.setFlags(age_item.flags() | Qt.ItemIsEditable)
                self.patient_table.setItem(row, 1, age_item)

                data_button = QPushButton("Data")
                data_button.clicked.connect(lambda _, uuid_key=uuid_key: self.open_data_screen(uuid_key))
                self.patient_table.setCellWidget(row, 2, data_button)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Error populating patient row: {str(e)}")
        self.patient_table.blockSignals(False)

    def update_patient_data(self, row, column):
        """Update the patient data based on table edits."""
        try:
            patient_uuid = list(self.patients.keys())[row]
            if column == 0:  # Name column
                self.patients[patient_uuid]["name"] = self.patient_table.item(row, column).text().strip()
                if not self.patients[patient_uuid]["name"]:
                    raise ValueError("Name cannot be empty.")
            elif column == 1:  # Age column
                age = int(self.patient_table.item(row, column).text().strip())
                if age <= 0:
                    raise ValueError("Age must be a positive number.")
                self.patients[patient_uuid]["age"] = age
            self.save_patients()
        except ValueError as e:
            QMessageBox.warning(self, "Invalid Input", str(e))
            self.populate_table()  # Revert invalid input

    def add_patient(self):
        """Add a new patient."""
        new_patient_uuid = str(uuid.uuid4())  # Generate a unique UUID for the new patient
        new_patient = {
            "name": "New Patient",
            "age": 30,
            "records": {}
        }
        self.patients[new_patient_uuid] = new_patient
        self.save_patients()
        self.populate_table()

    def save_questions_from_settings(self, updated_questions):
        """Save updated questions from the settings screen."""
        self.questions = updated_questions  # Update the in-memory list of questions
        self.save_questions()  # Save the questions to the JSON file

    def save_questions(self):
        with open(get_user_data_path("questions.json"), "w") as file:
            json.dump(self.questions, file, indent=4)

    def delete_patient(self):
        """Delete the selected patient."""
        selected_row = self.patient_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "No Selection", "Please select a patient to delete.")
            return

        # Confirm deletion
        patient_uuid = list(self.patients.keys())[selected_row]
        patient_name = self.patients[patient_uuid]["name"]
        confirmation = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete patient '{patient_name}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirmation == QMessageBox.No:
            return

        del self.patients[patient_uuid]
        self.save_patients()
        self.populate_table()

    def load_questions(self):
        path = get_user_data_path("questions.json")
        if not os.path.exists(path):
            default_questions = [
                {"text": "Question 1", "type": "Quantitative"},
                {"text": "Question 2", "type": "Qualitative"},
            ]
            with open(path, "w") as file:
                json.dump(default_questions, file, indent=4)
        with open(path, "r") as file:
            return json.load(file)

    def open_data_screen(self, patient_uuid):
        """Open the Data screen for the selected patient."""
        if patient_uuid in self.patients:
            self.data_window = DataScreen(patient_uuid, self.questions, self.patients)
            self.data_window.show()
        else:
            QMessageBox.warning(self, "Error", "Patient UUID not found.")

    def open_edit_data_screen(self):
        """Open the Edit Data screen."""
        self.edit_data_window = EditDataScreen(
            self.questions,
            self.save_questions_from_settings  # Pass the method as a callback
        )
        self.edit_data_window.show()


class DataScreen(QWidget):
    def __init__(self, patient_uuid, questions, patients):
        super().__init__()
        self.patient_uuid = patient_uuid
        self.patient = patients[self.patient_uuid]  # Retrieve patient data using UUID
        self.questions = questions
        self.setWindowTitle(f"Data for {self.patient['name']}")
        self.setGeometry(100, 100, 800, 600)

        # Load existing patient data
        self.patient_data = self.load_patient_data()

        # Initialize current date range (Monday to Friday)
        today = QDate.currentDate()
        self.start_date = self.get_week_start_date(today)
        self.end_date = self.start_date.addDays(4)

        # Build UI layout
        self.build_ui()
        self.populate_table()
        self.update_chart()

    def build_ui(self):
        """Build the UI layout for the data screen."""
        layout = QVBoxLayout()

        # Date Range Input
        date_range_layout = QHBoxLayout()
        self.start_date_label = QLabel(self.start_date.toString("yyyy-MM-dd"))
        self.end_date_label = QLabel(self.end_date.toString("yyyy-MM-dd"))
        prev_week_button = QPushButton("Previous Week")
        next_week_button = QPushButton("Next Week")
        prev_week_button.clicked.connect(self.go_to_previous_week)
        next_week_button.clicked.connect(self.go_to_next_week)
        date_range_layout.addWidget(QLabel("Start Date:"))
        date_range_layout.addWidget(self.start_date_label)
        date_range_layout.addWidget(QLabel("End Date:"))
        date_range_layout.addWidget(self.end_date_label)
        date_range_layout.addWidget(prev_week_button)
        date_range_layout.addWidget(next_week_button)
        layout.addLayout(date_range_layout)

        # Table for weekly data input
        self.data_table = QTableWidget()
        self.data_table.setColumnCount(6)  # Question + 5 days (Monday to Friday)
        self.data_table.setHorizontalHeaderLabels(["Question", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])
        self.populate_table()
        layout.addWidget(self.data_table)

        # Connect cellChanged to handle real-time updates
        self.data_table.cellChanged.connect(self.handle_table_edit)

        # Chart for quantitative data
        self.chart = QChart()
        self.chart_view = QChartView(self.chart)
        layout.addWidget(self.chart_view)

        # Buttons
        export_button = QPushButton("Export to Excel")
        export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(export_button)

        export_to_pdf_button = QPushButton("Export to PDF")
        export_to_pdf_button.clicked.connect(self.export_to_pdf)
        layout.addWidget(export_to_pdf_button)

        self.setLayout(layout)
        self.update_chart()

    def get_week_start_date(self, current_date):
        """Get the Monday of the current week based on the given date."""
        # Determine the day of the week (1 = Monday, 7 = Sunday)
        day_of_week = current_date.dayOfWeek()
        if day_of_week == 7:  # Adjust for Sunday being the last day of the week
            day_of_week = 0
        # Subtract days to get to Monday
        return current_date.addDays(-(day_of_week - 1))

    def handle_table_edit(self, row, column):
        """Handle edits to the data table and update the chart."""
        if column == 0:
            return  # Ignore edits to the "Question" column

        self.is_data_changed = True  # Mark as changed
        try:
            week_key = f"{self.start_date.toString('yyyy-MM-dd')}_to_{self.end_date.toString('yyyy-MM-dd')}"
            question = self.data_table.item(row, 0).text()
            day = DAYS_OF_WEEK[column - 1]
            value = self.data_table.item(row, column).text()

            # Update the patient data in memory
            week_data = self.patient_data.get(week_key, {})
            question_data = week_data.get(question, {})
            question_data[day] = value
            week_data[question] = question_data
            self.patient_data[week_key] = week_data

            # Autosave
            self.save_patient_data()  # Autosave changes

            # Update the chart dynamically
            self.update_chart()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to update data: {e}")

    def go_to_previous_week(self):
        """Navigate to the previous week's data."""
        self.start_date = self.start_date.addDays(-7)
        self.end_date = self.end_date.addDays(-7)
        self.update_date_labels()
        self.populate_table()
        self.update_chart()

    def go_to_next_week(self):
        """Navigate to the next week's data."""
        self.start_date = self.start_date.addDays(7)
        self.end_date = self.end_date.addDays(7)
        self.update_date_labels()
        self.populate_table()
        self.update_chart()

    def update_date_labels(self):
        """Update the displayed date range labels."""
        self.start_date_label.setText(self.start_date.toString("yyyy-MM-dd"))
        self.end_date_label.setText(self.end_date.toString("yyyy-MM-dd"))

    def populate_table(self):
        """Populate the data table with weekly inputs."""
        self.data_table.blockSignals(True)  # Prevent triggering cellChanged while populating
        self.data_table.setRowCount(0)
        week_key = f"{self.start_date.toString('yyyy-MM-dd')}_to_{self.end_date.toString('yyyy-MM-dd')}"

        # Get data for the selected week
        week_data = self.patient_data.get(week_key, {})
        for i, question in enumerate(self.questions):
            self.data_table.insertRow(i)

            # Question Name (non-editable)
            question_item = QTableWidgetItem(question["text"])
            question_item.setFlags(Qt.ItemIsEnabled)  # Make it non-editable
            self.data_table.setItem(i, 0, question_item)

            # Weekly data inputs (editable)
            for j, day in enumerate(DAYS_OF_WEEK, start=1):
                value = week_data.get(question["text"], {}).get(day, "")
                day_item = QTableWidgetItem(str(value))
                self.data_table.setItem(i, j, day_item)

        self.data_table.blockSignals(False)  # Re-enable cellChanged

    def update_chart(self):
        """Update the line chart with quantitative data for the current week."""
        self.chart.removeAllSeries()
        week_key = f"{self.start_date.toString('yyyy-MM-dd')}_to_{self.end_date.toString('yyyy-MM-dd')}"
        week_data = self.patient_data.get(week_key, {})

        for question in self.questions:
            if question["type"] == "Quantitative":
                series = QLineSeries()
                series.setName(question["text"])

                # Add data points for Monday to Friday
                question_data = week_data.get(question["text"], {})
                for i, day in enumerate(DAYS_OF_WEEK):
                    try:
                        value = float(question_data.get(day, 0) or 0)
                    except ValueError:
                        value = 0
                    series.append(i, value)

                self.chart.addSeries(series)

        # Configure the chart
        self.chart.createDefaultAxes()
        self.chart.setTitle(
            f"Quantitative Data ({self.start_date.toString('yyyy-MM-dd')} to {self.end_date.toString('yyyy-MM-dd')})")
        self.chart.legend().setVisible(True)

    def load_patient_data(self):
        """Load existing patient data for the specific patient."""
        try:
            path = get_user_data_path("patient_data.json")
            if os.path.exists(path):
                with open(path, "r") as file:
                    all_data = json.load(file)
                    return all_data.get(self.patient_uuid, {})  # Load only this patient's data
            else:
                return {}
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load patient data: {e}")
            return {}

    def save_patient_data(self):
        """Save the patient's data to a persistent file."""
        try:
            path = get_user_data_path("patient_data.json")
            all_data = {}
            if os.path.exists(path):
                with open(path, "r") as file:
                    all_data = json.load(file)
            all_data[self.patient_uuid] = self.patient_data
            with open(path, "w") as file:
                json.dump(all_data, file, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save patient data: {e}")

    def export_to_excel(self):
        """Export patient data to an Excel file with date range in the file name."""
        try:
            start_date_str = self.patient_data.get("start_date", "start")
            end_date_str = self.patient_data.get("end_date", "end")
            file_name = f"{self.patient['name']}_{start_date_str}_to_{end_date_str}_data.xlsx"
            save_path, _ = QFileDialog.getSaveFileName(
                self, "Save File", file_name, "Excel Files (*.xlsx)"
            )
            if not save_path:
                return

            # Create a temporary directory for the chart image
            temp_dir = tempfile.gettempdir()
            chart_image_path = os.path.join(temp_dir, f"{self.patient['name']}_chart.png")

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Data for {self.patient['name']}"

            ws.append(["Patient Name:", self.patient["name"]])
            ws.append(["Patient Age:", self.patient["age"]])
            ws.append(["Date Range:", f"{start_date_str} to {end_date_str}"])
            ws.append([])

            # Add headers
            ws.append(["Question"] + DAYS_OF_WEEK)

            for question, weekly_data in self.patient_data.items():
                if question in ["start_date", "end_date"]:  # Skip date metadata
                    continue
                row = [question]
                for day in DAYS_OF_WEEK:
                    row.append(weekly_data.get(day, ""))
                ws.append(row)

            # Add chart image
            try:
                chart_pixmap = self.chart_view.grab()
                chart_pixmap.save(chart_image_path)
                img = Image(chart_image_path)
                ws.add_image(img, "H2")
            except Exception as e:
                QMessageBox.warning(self, "Warning", f"Failed to add chart image: {str(e)}")

            # Save the Excel file
            wb.save(save_path)
            QMessageBox.information(self, "Exported", f"Data exported to {save_path}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export data: {str(e)}")
        finally:
            # Clean up temporary files
            if os.path.exists(chart_image_path):
                os.remove(chart_image_path)

    def export_to_pdf(self):
        """Export patient data to a PDF file."""
        try:
            start_date_str = self.start_date.toString("yyyy-MM-dd")
            end_date_str = self.end_date.toString("yyyy-MM-dd")
            file_name = f"{self.patient['name']}_{start_date_str}_to_{end_date_str}_data.pdf"
            save_path, _ = QFileDialog.getSaveFileName(
                self, "Save PDF", file_name, "PDF Files (*.pdf)"
            )
            if not save_path:
                return

            pdf = FPDF(orientation="P", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("Arial", size=12)

            # Add patient information
            pdf.set_font("Arial", style="B", size=14)
            pdf.cell(0, 10, f"Patient Name: {self.patient['name']}", ln=True)
            pdf.cell(0, 10, f"Patient Age: {self.patient['age']}", ln=True)
            pdf.cell(0, 10, f"Date Range: {start_date_str} to {end_date_str}", ln=True)
            pdf.ln(10)

            # Add table headers
            pdf.set_font("Arial", style="B", size=12)
            column_widths = [50, 25, 25, 25, 25, 25]  # Adjust to fit A4 size
            pdf.cell(column_widths[0], 10, "Question", border=1)
            for day in DAYS_OF_WEEK:
                pdf.cell(column_widths[1], 10, day, border=1)
            pdf.ln()

            # Add table data
            pdf.set_font("Arial", size=12)
            week_key = f"{start_date_str}_to_{end_date_str}"
            week_data = self.patient_data.get(week_key, {})

            for question in self.questions:
                pdf.cell(column_widths[0], 10, question["text"], border=1)
                for day in DAYS_OF_WEEK:
                    value = week_data.get(question["text"], {}).get(day, "")
                    pdf.cell(column_widths[1], 10, str(value), border=1)
                pdf.ln()

            # Add chart image
            temp_dir = tempfile.gettempdir()
            chart_image_path = os.path.join(temp_dir, "chart.png")
            try:
                chart_pixmap = self.chart_view.grab()
                chart_pixmap.save(chart_image_path)
                pdf.ln(10)  # Add spacing before the chart
                pdf.image(chart_image_path, x=10, y=pdf.get_y(), w=180)  # Fit chart within page width
            except Exception as e:
                QMessageBox.warning(self, "Warning", f"Failed to add chart image: {str(e)}")
            finally:
                if os.path.exists(chart_image_path):
                    os.remove(chart_image_path)

            # Save the PDF
            pdf.output(save_path)
            QMessageBox.information(self, "Exported", f"Data exported to {save_path}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export data to PDF: {str(e)}")


class EditDataScreen(QWidget):
    def __init__(self, questions, save_questions_callback):
        super().__init__()
        self.setWindowTitle("Edit Data Points")
        self.setGeometry(100, 100, 600, 400)

        # Store references
        self.questions = questions
        self.save_questions_callback = save_questions_callback

        # Main layout
        layout = QVBoxLayout()

        # Table for displaying questions
        self.question_table = QTableWidget()
        self.question_table.setColumnCount(2)
        self.question_table.setHorizontalHeaderLabels(["Question", "Type"])
        self.populate_table()
        layout.addWidget(self.question_table)

        # Buttons for adding and deleting questions
        button_layout = QHBoxLayout()
        add_button = QPushButton("Add")
        add_button.clicked.connect(self.add_question)
        button_layout.addWidget(add_button)

        delete_button = QPushButton("Delete")
        delete_button.clicked.connect(self.delete_question)
        button_layout.addWidget(delete_button)

        layout.addLayout(button_layout)

        # Save button
        save_button = QPushButton("Save")
        save_button.clicked.connect(self.save_questions)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def populate_table(self):
        """Populate the table with current questions."""
        self.question_table.setRowCount(0)
        for i, question in enumerate(self.questions):
            self.question_table.insertRow(i)

            # Question text (editable)
            question_item = QTableWidgetItem(question["text"])
            question_item.setFlags(question_item.flags() | Qt.ItemIsEditable)
            self.question_table.setItem(i, 0, question_item)

            # Checkbox for type
            checkbox = QCheckBox("Quantitative")
            checkbox.setChecked(question["type"] == "Quantitative")
            checkbox.setToolTip("Check for Quantitative, Uncheck for Qualitative")
            self.question_table.setCellWidget(i, 1, checkbox)

    def add_question(self):
        """Add a new question."""
        new_question = {"text": "New Question", "type": "Quantitative"}
        self.questions.append(new_question)
        self.populate_table()

    def delete_question(self):
        """Delete the selected question."""
        selected_row = self.question_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "No Selection", "Please select a question to delete.")
            return

        question_text = self.question_table.item(selected_row, 0).text()
        confirmation = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete the question '{question_text}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirmation == QMessageBox.Yes:
            del self.questions[selected_row]
            self.populate_table()

    def save_questions(self):
        """Save the changes to the questions."""
        # Update questions list from the table
        self.questions.clear()
        for row in range(self.question_table.rowCount()):
            question_text = self.question_table.item(row, 0).text()
            checkbox = self.question_table.cellWidget(row, 1)
            question_type = "Quantitative" if checkbox.isChecked() else "Qualitative"
            self.questions.append({"text": question_text, "type": question_type})

        # Save updated questions using the callback
        self.save_questions_callback(self.questions)
        QMessageBox.information(self, "Saved", "Data points have been updated successfully.")
        self.close()


if __name__ == "__main__":
    try:
        logging.info("Application started.")
        app = QApplication(sys.argv)

        # Initialize main application window
        window = EMRManager()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        logging.error(f"Application error: {e}")